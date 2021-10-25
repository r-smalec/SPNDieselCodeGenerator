from openpyxl import load_workbook

def get_SPN(excelLine):
    if ws.cell(excelLine, 7).value is not None:
        return ws.cell(excelLine, 7).value
    else:
        return 0

def get_FMI(excelLine):
    if ws.cell(excelLine, 8).value is not None:
        return ws.cell(excelLine, 8).value
    else:
        return 255

def get_uiCodeNb(excelLine):
    if ws.cell(excelLine, 1).value is not None:
        return str(ws.cell(excelLine, 1).value)
    else:
        return 0

def get_sPriority(excelLine):
    if ws.cell(excelLine, 3).value is not None:
        return str("'" + ws.cell(excelLine, 3).value + "'")
    else:
        return 0

def get_usType(excelLine):
    if ws.cell(excelLine, 4).value is not None:
        return str(ws.cell(excelLine, 4).value)
    else:
        return 0

def get_uiDeviceNb(excelLine):
    if ws.cell(excelLine, 5).value is not None:
        return str(ws.cell(excelLine, 5).value)
    else:
        return str("1")

def get_sDescription(excelLine):
    strange = 'ęóąśłżźćń'
    ascii_replacements = 'eoaslzzcn'
    translator = str.maketrans(strange,ascii_replacements)
    excelText = ws.cell(excelLine, 6).value
    if excelText is not None:
        return str("'" + excelText.translate(translator) + "'")
    else:
        return 0

def get_sDescriptionStaff():
    return str("''")

def get_sDescriptionMaintenance():
    return str("''")

def get_uiFaultNb(excelLine):
    if ws.cell(excelLine, 2).value is not None:
        return str (ws.cell(excelLine, 2).value)
    else:
       return 0

try:
    wb = load_workbook('Bledy_silniki_Spalinowe_D1.xlsx')
except:
    print("No such file")
    quit()

ws = wb.active
linesCount = 1
while(ws.cell(linesCount, 1).value is not None):
    linesCount += 1

try:
    codeFile = open('codeFile_D1.txt', 'w')
except:
    print("File could not be created")
    quit()

#print("(*SOFTCONTROL:\nVERSION:4.00.20*)\nFUNCTION_BLOCK DIAG_WyszukajKomunikatySilnika\n(**)\n(**)", file = codeFile)
print("\tVAR_INPUT", file = codeFile)
print("\t\tdi_SzukaneSPN: DINT:=0;", file = codeFile)
print("\t\tb_SzukaneFMI: BYTE:=0;", file = codeFile)
print("\tEND_VAR", file = codeFile)
print("\tVAR_OUTPUT", file = codeFile)
print("\t\tuiCodeNb: UINT:=0", file = codeFile)
print("\t\tsPriority: STRING[1]:=''", file = codeFile)
print("\t\tusType: USINT:=0", file = codeFile)
print("\t\tsDescription: STRING[255]:=''", file = codeFile)
print("\t\tuiFaultNumber: UINT:=0", file = codeFile)
print("\tEND_VAR", file = codeFile)
print("\tVAR", file = codeFile)
print("\t\tsDescriptionStaff: STRING[1]:=''", file = codeFile)
print("\t\tsDescriptionMaintenance: STRING[1]:=''", file = codeFile)
print("\t\tuiDeviceNb: UINT:=0", file = codeFile)
print("\tEND_VAR", file = codeFile)
print("'ST'", file = codeFile)
print("BODY", file = codeFile)

previousSPN = 0

line = 2
while(line < linesCount):
    currentSPN = get_SPN(line)
    nextSPN = get_SPN(line + 1)

    if line == 2:
        print("IF di_SzukaneSPN = " + str(currentSPN) + " THEN", file = codeFile)
    else:
        print("ELSIF di_SzukaneSPN = " + str(currentSPN) + " THEN", file = codeFile)

    if nextSPN is not currentSPN:
        caseStr = "\tuiCodeNb := " + get_uiCodeNb(line) + ";\n" + "\tsPriority := " + get_sPriority(line) + ";\n" +  "\tusType := " + get_usType(line) + ";\n" + "\tuiDeviceNb := " + get_uiDeviceNb(line) + ";\n" + "\tsDescription := " + get_sDescription(line) + ";\n" + "\tsDescriptionStaff := " + get_sDescriptionStaff() + ";\n" + "\tsDescriptionMaintenance := " + get_sDescriptionMaintenance() + ";\n" + "\tUiFaultNumber := " + get_uiFaultNb(line) + ";"
        print(caseStr, file = codeFile)
        line += 1
    else:
        newLine = line
        while((get_SPN(newLine) is get_SPN(newLine + 1)) or (get_SPN(newLine) is get_SPN(newLine - 1))):
            currentFMI = get_FMI(newLine)
            print("\tIF b_SzukaneFMI = " + str(currentFMI) + " THEN", file = codeFile)
            caseStr = "\t\tuiCodeNb := " + get_uiCodeNb(newLine) + ";\n" + "\t\tsPriority := " + get_sPriority(newLine) + ";\n" +  "\t\tusType := " + get_usType(newLine) + ";\n" + "\t\tuiDeviceNb := " + get_uiDeviceNb(newLine) + ";\n" + "\t\tsDescription := " + get_sDescription(newLine) + ";\n" + "\t\tsDescriptionStaff := " + get_sDescriptionStaff() + ";\n" + "\t\tsDescriptionMaintenance := " + get_sDescriptionMaintenance() + ";\n" + "\t\tUiFaultNumber := " + get_uiFaultNb(newLine) + ";"
            print(caseStr, file = codeFile)
            newLine += 1
            print("\tEND_IF;", file = codeFile)
        line = newLine
    previousSPN = get_SPN(line)
print("ELSE", file = codeFile)
caseStr = "\tuiCodeNb := 0;\n\tsPriority := 'A';\n\tusType := 0;\n\tuiDeviceNb := 1;\n\tsDescription := 'Niezdefiniowany blad silnika';\n\tsDescriptionStaff := '';\n\tsDescriptionMaintenance := '';\n\tUiFaultNumber := 0;"
print(caseStr, file = codeFile)
print("END_IF;\nEND_BODY\nEND_FUNCTION_BLOCK", file = codeFile)